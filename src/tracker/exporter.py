"""Export applications table to Excel with formatting."""

from __future__ import annotations

from pathlib import Path
from sqlite3 import Connection

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from tracker import db as dbmod


def gmail_thread_url(thread_id: str) -> str:
    return f"https://mail.google.com/mail/u/0/#all/{thread_id}"


def export_xlsx(conn: Connection, output_path: Path) -> None:
    apps = dbmod.list_applications(conn)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Applications"

    headers = [
        "Company",
        "Role",
        "Location",
        "Pay",
        "Status",
        "Date Applied",
        "Last Update",
        "Gmail Thread",
    ]
    ws.append(headers)

    for app in apps:
        url = gmail_thread_url(app.gmail_thread_id)
        ws.append(
            [
                app.company,
                app.role,
                app.location or "",
                app.pay or "",
                app.status,
                app.first_seen_at.strftime("%Y-%m-%d %H:%M"),
                app.last_updated_at.strftime("%Y-%m-%d %H:%M"),
                "",
            ]
        )
        r = ws.max_row
        cell = ws.cell(row=r, column=8)
        cell.hyperlink = url
        cell.value = "Open in Gmail"
        cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    fills = {
        "offer": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "interview": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "applied": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    }
    if last_row >= 2:
        for status, fill in fills.items():
            ws.conditional_formatting.add(
                f"E2:E{last_row}",
                CellIsRule(operator="equal", formula=[f'"{status}"'], fill=fill),
            )

    widths = [28, 36, 24, 18, 14, 18, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
